from __future__ import annotations

from datetime import date, datetime, timezone
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
import pandas as pd


TARGET_SHEET = "Revenue Forecast Q1&2 - 2026"
TARGET_DATE_COLUMN = "Date"
TARGET_PRIOR_COLUMN = "DAILY REV 2025"
TARGET_FORECAST_COLUMN = "FORECAST 2026"
TARGET_COLUMNS = ["Date", "Revenue 2025", "Forecast 2026"]
TARGET_SCHEMA_VERSION = "fbm-daily-target-v2"
_FORECAST_FORMULA = re.compile(r"=C(\d+)(?:\*([0-9.]+))?", re.IGNORECASE)


class TargetError(RuntimeError):
    pass


def empty_fbm_target() -> pd.DataFrame:
    return pd.DataFrame(columns=TARGET_COLUMNS)


def normalize_fbm_daily_target(frame: pd.DataFrame) -> pd.DataFrame:
    missing = [column for column in TARGET_COLUMNS if column not in frame]
    if missing:
        raise TargetError("Dữ liệu target ngày thiếu cột: " + ", ".join(missing))
    normalized = frame[TARGET_COLUMNS].copy()
    normalized["Date"] = pd.to_datetime(normalized["Date"], errors="coerce")
    for column in ("Revenue 2025", "Forecast 2026"):
        normalized[column] = pd.to_numeric(normalized[column], errors="coerce")
    normalized = normalized.dropna(subset=TARGET_COLUMNS).copy()
    if normalized.empty:
        raise TargetError("Không tìm thấy target ngày hợp lệ trong sheet đã chọn.")
    if normalized["Date"].duplicated().any():
        raise TargetError("Target ngày có Date bị trùng.")
    normalized["Date"] = normalized["Date"].dt.strftime("%Y-%m-%d")
    return normalized.sort_values("Date", kind="stable").reset_index(drop=True)


def _forecast_value(value: Any, row: int, calculated: dict[int, float]) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if not isinstance(value, str):
        raise TargetError(f"FORECAST 2026 tại dòng {row} không hợp lệ.")
    match = _FORECAST_FORMULA.fullmatch(value.replace(" ", ""))
    if not match:
        raise TargetError(f"Không đọc được công thức FORECAST 2026 tại dòng {row}: {value}")
    reference_row = int(match.group(1))
    if reference_row not in calculated:
        raise TargetError(f"FORECAST 2026 dòng {row} tham chiếu dòng chưa có dữ liệu.")
    factor = float(match.group(2)) if match.group(2) else 1.0
    return calculated[reference_row] * factor


def read_fbm_target_workbook(path: Path, *, year: int = 2026) -> pd.DataFrame:
    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
    except (OSError, ValueError, KeyError) as exc:
        raise TargetError("Không thể đọc file target FBM.") from exc
    try:
        if TARGET_SHEET not in workbook.sheetnames:
            raise TargetError(f"Không tìm thấy sheet `{TARGET_SHEET}`.")
        sheet = workbook[TARGET_SHEET]
        headers = [sheet.cell(1, column).value for column in range(1, 4)]
        expected = [TARGET_DATE_COLUMN, TARGET_PRIOR_COLUMN, TARGET_FORECAST_COLUMN]
        if headers != expected:
            raise TargetError(
                "Sheet target phải có đúng 3 cột đầu: " + ", ".join(expected)
            )

        # Cột Date nguồn đang trộn Excel serial và text do định dạng locale.
        # Sheet quy ước đúng một dòng cho mỗi ngày năm 2026, theo thứ tự từ
        # 01/01 đến 31/12, nên dùng vị trí dòng làm trục ngày ổn định.
        dates = pd.date_range(f"{year}-01-01", f"{year}-12-31", freq="D")
        calculated: dict[int, float] = {}
        rows: list[dict[str, Any]] = []
        for offset, current_date in enumerate(dates):
            row = offset + 2
            prior = pd.to_numeric(sheet.cell(row, 2).value, errors="coerce")
            if pd.isna(prior):
                raise TargetError(f"DAILY REV 2025 tại dòng {row} không hợp lệ.")
            forecast = _forecast_value(sheet.cell(row, 3).value, row, calculated)
            calculated[row] = forecast
            rows.append(
                {
                    "Date": current_date,
                    "Revenue 2025": float(prior),
                    "Forecast 2026": forecast,
                }
            )
        return normalize_fbm_daily_target(pd.DataFrame(rows))
    finally:
        workbook.close()


def target_metadata_path(path: Path) -> Path:
    return path.with_suffix(".metadata.json")


def save_fbm_target_snapshot(
    path: Path,
    frame: pd.DataFrame,
    *,
    source_name: str = "",
) -> None:
    normalized = normalize_fbm_daily_target(frame)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    normalized.to_csv(temporary, index=False, encoding="utf-8")
    temporary.replace(path)

    dates = pd.to_datetime(normalized["Date"], errors="coerce")
    metadata = {
        "schema_version": TARGET_SCHEMA_VERSION,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "source_name": source_name,
        "source_sheet": TARGET_SHEET,
        "source_columns": [TARGET_DATE_COLUMN, TARGET_PRIOR_COLUMN, TARGET_FORECAST_COLUMN],
        "scope": "All Stores · FBM",
        "date_min": dates.min().date().isoformat(),
        "date_max": dates.max().date().isoformat(),
        "rows": len(normalized),
    }
    sidecar = target_metadata_path(path)
    temporary_metadata = sidecar.with_suffix(sidecar.suffix + ".tmp")
    temporary_metadata.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    temporary_metadata.replace(sidecar)


def load_fbm_target_snapshot(path: Path) -> pd.DataFrame:
    if not path.exists():
        return empty_fbm_target()
    try:
        frame = pd.read_csv(path, dtype={"Date": str})
    except (OSError, pd.errors.ParserError, UnicodeDecodeError) as exc:
        raise TargetError("Không thể đọc snapshot target FBM.") from exc
    return normalize_fbm_daily_target(frame)


def load_fbm_target_metadata(path: Path) -> dict[str, Any]:
    sidecar = target_metadata_path(path)
    if not sidecar.exists():
        return {}
    try:
        return json.loads(sidecar.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        raise TargetError("Không thể đọc metadata target FBM.") from exc


def daily_targets_for_month(frame: pd.DataFrame, month: str) -> pd.DataFrame:
    if frame.empty:
        return empty_fbm_target()
    normalized = normalize_fbm_daily_target(frame)
    dates = pd.to_datetime(normalized["Date"], errors="coerce")
    result = normalized.loc[dates.dt.strftime("%Y-%m").eq(month)].copy()
    result["Date"] = pd.to_datetime(result["Date"], errors="coerce")
    return result.reset_index(drop=True)


def target_for_month(frame: pd.DataFrame, month: str) -> float | None:
    selected = daily_targets_for_month(frame, month)
    if selected.empty:
        return None
    return float(selected["Forecast 2026"].sum())


def target_progress(
    frame: pd.DataFrame,
    month: str,
    actual: float,
    as_of_date: date | pd.Timestamp | str,
) -> dict[str, float | int]:
    selected = daily_targets_for_month(frame, month)
    if selected.empty:
        raise TargetError(f"Không có target ngày cho tháng {month}.")
    month_start = pd.Timestamp(f"{month}-01")
    month_end = month_start + pd.offsets.MonthEnd(1)
    as_of = pd.Timestamp(as_of_date).normalize()
    cutoff = min(max(as_of, month_start - pd.Timedelta(days=1)), month_end)
    elapsed = selected[selected["Date"].le(cutoff)]
    forecast_mtd = float(elapsed["Forecast 2026"].sum())
    prior_mtd = float(elapsed["Revenue 2025"].sum())
    actual = float(actual)
    return {
        "elapsed_days": len(elapsed),
        "forecast_mtd": forecast_mtd,
        "prior_mtd": prior_mtd,
        "vs_forecast": actual / forecast_mtd - 1 if forecast_mtd else 0.0,
        "vs_2025": actual / prior_mtd - 1 if prior_mtd else 0.0,
        "yoy_index": actual / prior_mtd if prior_mtd else 0.0,
        "forecast_full_month": float(selected["Forecast 2026"].sum()),
        "prior_full_month": float(selected["Revenue 2025"].sum()),
    }
