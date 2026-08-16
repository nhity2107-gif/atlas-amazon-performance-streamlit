from __future__ import annotations

from datetime import date
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook
import pandas as pd

from target_data import (
    TARGET_SHEET,
    TargetError,
    daily_targets_for_month,
    load_fbm_target_snapshot,
    normalize_fbm_daily_target,
    read_fbm_target_workbook,
    save_fbm_target_snapshot,
    target_for_month,
    target_progress,
)


def write_forecast_workbook(path: Path) -> None:
    workbook = Workbook()
    output = workbook.active
    output.title = "Output Plan"
    output.append(["Date", "DAILY REV 2025", "FORECAST 2026"])
    output.append(["wrong", 999, 999])
    sheet = workbook.create_sheet(TARGET_SHEET)
    sheet.append(["Date", "DAILY REV 2025", "FORECAST 2026"])
    dates = pd.date_range("2026-01-01", "2026-12-31", freq="D")
    for index, current_date in enumerate(dates, start=2):
        forecast = 100 if index == 2 else f"=C{index - 1}"
        sheet.append([current_date.to_pydatetime(), 50, forecast])
    workbook.save(path)


class TargetDataTests(unittest.TestCase):
    def test_workbook_reads_only_daily_columns_from_named_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "forecast.xlsx"
            write_forecast_workbook(path)
            result = read_fbm_target_workbook(path)
        self.assertEqual(len(result), 365)
        self.assertEqual(result.iloc[0].to_dict(), {
            "Date": "2026-01-01",
            "Revenue 2025": 50.0,
            "Forecast 2026": 100.0,
        })
        self.assertEqual(target_for_month(result, "2026-08"), 3100.0)

    def test_snapshot_round_trip(self) -> None:
        source = normalize_fbm_daily_target(
            pd.DataFrame(
                {
                    "Date": ["2026-08-01", "2026-08-02"],
                    "Revenue 2025": [50, 60],
                    "Forecast 2026": [100, 120],
                }
            )
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "fbm_target.csv"
            save_fbm_target_snapshot(path, source, source_name="forecast.xlsx")
            result = load_fbm_target_snapshot(path)
        pd.testing.assert_frame_equal(result, source)

    def test_daily_progress_compares_actual_to_forecast_and_prior_year(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "forecast.xlsx"
            write_forecast_workbook(path)
            frame = read_fbm_target_workbook(path)
        result = target_progress(frame, "2026-08", 1200, date(2026, 8, 16))
        self.assertEqual(result["elapsed_days"], 16)
        self.assertEqual(result["forecast_mtd"], 1600)
        self.assertEqual(result["prior_mtd"], 800)
        self.assertAlmostEqual(result["vs_forecast"], -0.25)
        self.assertAlmostEqual(result["vs_2025"], 0.5)
        self.assertEqual(result["forecast_full_month"], 3100)

    def test_completed_month_uses_every_daily_value(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "forecast.xlsx"
            write_forecast_workbook(path)
            frame = read_fbm_target_workbook(path)
        result = target_progress(frame, "2026-07", 2000, date(2026, 8, 16))
        self.assertEqual(result["elapsed_days"], 31)
        self.assertEqual(result["forecast_mtd"], 3100)
        self.assertEqual(len(daily_targets_for_month(frame, "2026-07")), 31)

    def test_rejects_wrong_headers(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "wrong.xlsx"
            workbook = Workbook()
            workbook.active.title = TARGET_SHEET
            workbook.active.append(["Month", "Other", "Forecast"])
            workbook.save(path)
            with self.assertRaises(TargetError):
                read_fbm_target_workbook(path)


if __name__ == "__main__":
    unittest.main()
